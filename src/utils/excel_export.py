"""Xuất thống kê kèo ra file Excel (.xlsx)."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── màu sắc ──────────────────────────────────────────────────────────────
_GREEN   = "FF198754"
_RED     = "FFDC3545"
_YELLOW  = "FFFFC107"
_BLUE    = "FF0D6EFD"
_GRAY    = "FF6C757D"
_HEADER  = "FF212529"
_BG_WIN  = "FFD1FAE5"
_BG_LOSS = "FFFFE4E6"
_BG_EVEN = "FFF8F9FA"
_WHITE   = "FFFFFFFF"

_thin = Side(style="thin", color="FFD1D5DB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(cell, text, bold=True, bg=_HEADER, fg=_WHITE, wrap=False):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _border


def _cell(cell, value, bold=False, color=None, align="left", bg=None, fmt=None):
    cell.value = value
    cell.font = Font(bold=bold, color=color or "FF000000", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt


def _parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


def _exit_price(s: dict) -> float | None:
    status = s["status"]
    if s.get("close_price"):
        return s["close_price"]
    if status == "SL_HIT":
        return s["sl_price"]
    if status == "TP1_HIT":
        return s["tp1"]
    if status == "TP2_HIT":
        return s["tp2"] or s["tp1"]
    if status == "TP3_HIT":
        return s["tp3"] or s["tp2"] or s["tp1"]
    return None


def _pct(s: dict) -> float | None:
    entry = s["entry_price"]
    exit_px = _exit_price(s)
    if exit_px is None or not entry:
        return None
    if s["direction"] == "LONG":
        return (exit_px - entry) / entry * 100
    return (entry - exit_px) / entry * 100


def _duration_min(s: dict) -> int | None:
    t0 = _parse_dt(s.get("created_at"))
    t1 = _parse_dt(s.get("closed_at"))
    if t0 and t1:
        return max(0, int((t1 - t0).total_seconds() / 60))
    return None


def _is_closed(s: dict) -> bool:
    return s["status"] in ("TP1_HIT", "TP2_HIT", "TP3_HIT", "SL_HIT")


def _status_label(status: str) -> str:
    return {
        "PENDING":  "⏳ Pending",
        "ACTIVE":   "🟡 Active",
        "TP1_HIT":  "✅ TP1",
        "TP2_HIT":  "✅ TP2",
        "TP3_HIT":  "✅ TP3",
        "SL_HIT":   "❌ SL",
        "CANCELLED":"⚪ Cancelled",
    }.get(status, status)


# ── Sheet 1: Tổng quan ───────────────────────────────────────────────────
def _write_summary(wb: openpyxl.Workbook, signals: list[dict]):
    ws = wb.create_sheet("Tổng quan")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    closed = [s for s in signals if _is_closed(s)]
    wins   = [s for s in closed if s["status"] != "SL_HIT"]
    losses = [s for s in closed if s["status"] == "SL_HIT"]
    active = [s for s in signals if s["status"] in ("ACTIVE", "PENDING")]
    cancelled = [s for s in signals if s["status"] == "CANCELLED"]

    win_rate = len(wins) / len(closed) * 100 if closed else 0

    pcts_win  = [p for s in wins   if (p := _pct(s)) is not None]
    pcts_loss = [p for s in losses if (p := _pct(s)) is not None]
    avg_win  = sum(pcts_win)  / len(pcts_win)  if pcts_win  else 0
    avg_loss = sum(pcts_loss) / len(pcts_loss) if pcts_loss else 0
    expectancy = (win_rate / 100) * avg_win + (1 - win_rate / 100) * avg_loss

    durs_win  = [d for s in wins   if (d := _duration_min(s)) is not None]
    durs_loss = [d for s in losses if (d := _duration_min(s)) is not None]
    avg_dur_win  = sum(durs_win)  // len(durs_win)  if durs_win  else 0
    avg_dur_loss = sum(durs_loss) // len(durs_loss) if durs_loss else 0

    tp1 = sum(1 for s in wins if s["status"] == "TP1_HIT")
    tp2 = sum(1 for s in wins if s["status"] == "TP2_HIT")
    tp3 = sum(1 for s in wins if s["status"] == "TP3_HIT")

    rows = [
        ("📌 Tổng kèo",         len(signals),         None),
        ("✅ Thắng",             len(wins),            _GREEN),
        ("❌ Thua",              len(losses),          _RED),
        ("🟡 Active / Pending", len(active),          _YELLOW),
        ("⚪ Đã hủy",           len(cancelled),       _GRAY),
        (None, None, None),
        ("📊 Win rate",          f"{win_rate:.1f}%",  _GREEN if win_rate >= 50 else _RED),
        ("📈 Avg % thắng",       f"+{avg_win:.2f}%",  _GREEN),
        ("📉 Avg % thua",        f"{avg_loss:.2f}%",  _RED),
        ("💡 Expectancy",        f"{'+' if expectancy >= 0 else ''}{expectancy:.2f}%",
                                 _GREEN if expectancy >= 0 else _RED),
        (None, None, None),
        ("⏱ Avg thời gian thắng", f"{avg_dur_win // 60}h{avg_dur_win % 60:02d}m", None),
        ("⏱ Avg thời gian thua",  f"{avg_dur_loss // 60}h{avg_dur_loss % 60:02d}m", None),
        (None, None, None),
        ("🎯 TP1 hit",           tp1,  None),
        ("🎯 TP2 hit",           tp2,  None),
        ("🎯 TP3 hit",           tp3,  None),
    ]

    # header row
    ws.merge_cells("A1:B1")
    _hdr(ws["A1"], "📊 Tổng hợp kèo", bg="FF0D6EFD")
    ws.row_dimensions[1].height = 24

    for i, (label, value, color) in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 20
        if label is None:
            continue
        _cell(ws.cell(i, 1), label, bold=True, align="left")
        c = ws.cell(i, 2)
        _cell(c, value, bold=True, align="center",
              color=color, bg=_BG_EVEN if i % 2 == 0 else _WHITE)


# ── Sheet 2: Tất cả kèo ──────────────────────────────────────────────────
_TRADE_COLS = [
    ("ID",        7),
    ("Coin",      8),
    ("Direction", 9),
    ("Entry",     12),
    ("TP1",       12),
    ("TP2",       12),
    ("TP3",       12),
    ("SL",        12),
    ("Exit Price",12),
    ("% P/L",     10),
    ("Status",    13),
    ("Source",    12),
    ("Leverage",   9),
    ("Quality",    9),
    ("Order Type", 11),
    ("Note",       22),
    ("Tạo lúc",   18),
    ("Đóng lúc",  18),
    ("Thời gian", 11),
    ("Close Reason", 14),
]


def _write_trades_sheet(wb: openpyxl.Workbook, signals: list[dict], title: str, filter_fn=None):
    ws = wb.create_sheet(title)
    data = [s for s in signals if filter_fn is None or filter_fn(s)]

    for ci, (col_name, width) in enumerate(_TRADE_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        _hdr(ws.cell(1, ci), col_name)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, s in enumerate(data, start=2):
        is_win  = s["status"] in ("TP1_HIT", "TP2_HIT", "TP3_HIT")
        is_loss = s["status"] == "SL_HIT"
        row_bg  = _BG_WIN if is_win else (_BG_LOSS if is_loss else _WHITE)

        dur = _duration_min(s)
        dur_str = ""
        if dur is not None:
            h, m = divmod(dur, 60)
            dur_str = f"{h}h{m:02d}m" if h else f"{m}m"

        pct = _pct(s)
        exit_px = _exit_price(s)

        vals = [
            s["id"],
            s["coin"],
            s["direction"],
            s["entry_price"],
            s.get("tp1"),
            s.get("tp2"),
            s.get("tp3"),
            s["sl_price"],
            exit_px,
            pct,
            _status_label(s["status"]),
            s.get("source", ""),
            s.get("leverage"),
            s.get("quality_score"),
            s.get("order_type", ""),
            s.get("note") or "",
            s.get("created_at", "")[:16] if s.get("created_at") else "",
            s.get("closed_at", "")[:16]  if s.get("closed_at")  else "",
            dur_str,
            s.get("close_reason") or "",
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci)
            fmt = None
            align = "left"
            bold = False
            color = None

            col_name = _TRADE_COLS[ci - 1][0]
            if col_name in ("Entry", "TP1", "TP2", "TP3", "SL", "Exit Price"):
                fmt = '#,##0.########'
                align = "right"
            elif col_name == "% P/L":
                fmt = '+0.00%;-0.00%'
                align = "right"
                bold = True
                if pct is not None:
                    color = _GREEN if pct >= 0 else _RED
                if val is not None:
                    val = val / 100
            elif col_name == "ID":
                align = "center"
            elif col_name in ("Direction", "Status", "Source", "Order Type", "Leverage", "Quality"):
                align = "center"

            _cell(c, val, bold=bold, color=color, align=align, bg=row_bg, fmt=fmt)

        ws.row_dimensions[ri].height = 18

    # auto-filter
    if data:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_TRADE_COLS))}{len(data) + 1}"


# ── Sheet 5: Theo nguồn ──────────────────────────────────────────────────
def _write_by_source(wb: openpyxl.Workbook, signals: list[dict]):
    ws = wb.create_sheet("Theo nguồn")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    headers = ["Source", "Thắng", "Thua", "Tổng", "Win rate", "Avg % Win", "Avg % Loss"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws.cell(1, ci), h)
    ws.row_dimensions[1].height = 22

    # group by source
    sources: dict[str, list] = {}
    for s in signals:
        if not _is_closed(s):
            continue
        src = s.get("source") or "UNKNOWN"
        sources.setdefault(src, []).append(s)

    for ri, (src, rows) in enumerate(sorted(sources.items()), start=2):
        wins   = [s for s in rows if s["status"] != "SL_HIT"]
        losses = [s for s in rows if s["status"] == "SL_HIT"]
        wr = len(wins) / len(rows) * 100 if rows else 0
        pcts_w = [p for s in wins   if (p := _pct(s)) is not None]
        pcts_l = [p for s in losses if (p := _pct(s)) is not None]
        avg_w = sum(pcts_w) / len(pcts_w) if pcts_w else 0
        avg_l = sum(pcts_l) / len(pcts_l) if pcts_l else 0

        bg = _BG_EVEN if ri % 2 == 0 else _WHITE
        _cell(ws.cell(ri, 1), src,            bold=True, align="center", bg=bg)
        _cell(ws.cell(ri, 2), len(wins),      align="center", bg=bg, color=_GREEN)
        _cell(ws.cell(ri, 3), len(losses),    align="center", bg=bg, color=_RED)
        _cell(ws.cell(ri, 4), len(rows),      align="center", bg=bg)
        c_wr = ws.cell(ri, 5)
        _cell(c_wr, wr / 100, align="center", bold=True,
              color=_GREEN if wr >= 50 else _RED, bg=bg, fmt="0.0%")
        _cell(ws.cell(ri, 6), avg_w / 100,   align="center", color=_GREEN, bg=bg, fmt="+0.00%;-0.00%")
        _cell(ws.cell(ri, 7), avg_l / 100,   align="center", color=_RED,   bg=bg, fmt="+0.00%;-0.00%")
        ws.row_dimensions[ri].height = 18


# ── Sheet: Giả định P&L ──────────────────────────────────────────────────
_CLOSED_SIM = {"TP1_HIT", "TP2_HIT", "TP3_HIT", "SL_HIT", "CANCELLED"}

def _write_sim_sheet(wb: openpyxl.Workbook, signals: list[dict], sim_vol: float):
    ws = wb.create_sheet("Giả định P&L")
    headers = [
        ("ID", 7), ("Coin", 8), ("Dir", 7), ("Entry", 12),
        ("TP1", 12), ("TP2", 12), ("TP3", 12), ("SL", 12),
        ("Exit Price", 12), ("Status", 13),
        (f"Sim P&L (${sim_vol:.0f})", 16), ("Tạo lúc", 18),
    ]
    for ci, (h, w) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _hdr(ws.cell(1, ci), h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    closed = [s for s in signals if s["status"] in _CLOSED_SIM]
    total_sim = 0.0

    for ri, s in enumerate(closed, start=2):
        sim_pnl = s.get("sim_pnl")
        is_win  = s["status"] in ("TP1_HIT", "TP2_HIT", "TP3_HIT")
        is_loss = s["status"] == "SL_HIT"
        row_bg  = _BG_WIN if is_win else (_BG_LOSS if is_loss else _WHITE)
        exit_px = _exit_price(s)

        vals = [
            s["id"], s["coin"], s["direction"],
            s["entry_price"], s.get("tp1"), s.get("tp2"), s.get("tp3"), s["sl_price"],
            exit_px, _status_label(s["status"]),
            sim_pnl,
            s.get("created_at", "")[:16] if s.get("created_at") else "",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci)
            h_name = headers[ci - 1][0]
            fmt, align, bold, color = None, "left", False, None
            if h_name in ("Entry", "TP1", "TP2", "TP3", "SL", "Exit Price"):
                fmt, align = '#,##0.########', "right"
            elif "Sim P&L" in h_name:
                fmt, align, bold = '#,##0.00;[Red]-#,##0.00', "right", True
                if sim_pnl is not None:
                    color = _GREEN if sim_pnl >= 0 else _RED
                if sim_pnl is not None:
                    total_sim += sim_pnl
            elif h_name in ("ID", "Dir", "Status"):
                align = "center"
            _cell(c, val, bold=bold, color=color, align=align, bg=row_bg, fmt=fmt)
        ws.row_dimensions[ri].height = 18

    # Tổng cộng
    total_row = len(closed) + 2
    ws.row_dimensions[total_row].height = 22
    _cell(ws.cell(total_row, 10), "TỔNG", bold=True, align="right")
    c_total = ws.cell(total_row, 11)
    _cell(c_total, total_sim, bold=True, align="right",
          color=_GREEN if total_sim >= 0 else _RED,
          bg=_BG_WIN if total_sim >= 0 else _BG_LOSS,
          fmt='#,##0.00;[Red]-#,##0.00')

    if closed:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(closed) + 1}"


# ── Entry point ──────────────────────────────────────────────────────────
def build_excel(signals: list[dict], sim_vol: float | None = None) -> bytes:
    """Tạo workbook Excel từ danh sách signals, trả về bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định

    _write_summary(wb, signals)
    _write_trades_sheet(wb, signals, "Tất cả kèo")
    _write_trades_sheet(wb, signals, "Kèo thắng",
                        filter_fn=lambda s: s["status"] in ("TP1_HIT", "TP2_HIT", "TP3_HIT"))
    _write_trades_sheet(wb, signals, "Kèo thua",
                        filter_fn=lambda s: s["status"] == "SL_HIT")
    _write_by_source(wb, signals)
    if sim_vol is not None:
        _write_sim_sheet(wb, signals, sim_vol)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
